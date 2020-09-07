from openpyxl import workbook
from openpyxl import load_workbook

class Save_File:
    def __init__(self):
        self.local_path = ''
        self.online_path = ''

    def check_for_local_saved_data(self):
        wb = load_workbook(self.local_path)
        sheets = wb.sheetnames
        Game_State = wb[sheets[2]]
        if Game_State.cell(row = 9, column = 2).value:
            return True
        else:
            return False

    def check_for_online_saved_data(self):
        wb = load_workbook(self.online_path)
        sheets = wb.sheetnames
        Game_State = wb[sheets[2]]
        if Game_State.cell(row = 9, column = 2).value == True:
            return True
        else:
            return False

    def save_game(self, Player, Enemy, Game):
        if Game.Online == True:
            wb = load_workbook(self.online_path)
        else:
            wb = load_workbook(self.local_path)
        sheets = wb.sheetnames
        Player1 = wb[sheets[0]]
        Player2 = wb[sheets[1]]
        Game_State = wb[sheets[2]]
        #Clear Save File
        for row in Player1['B2:C18']:
            for cell in row:
                cell.value = None
        for row in Player1['A19:C119']:
            for cell in row:
                cell.value = None
        for row in Player2['B2:C18']:
            for cell in row:
                cell.value = None
        for row in Player2['A19:C119']:
            for cell in row:
                cell.value = None
        for row in Game_State['B2:B9']:
            for cell in row:
                cell.value = None
        #Save Player Names
        Player1.cell(row = 1, column = 1).value = Player.Name
        Player2.cell(row = 1, column = 1).value = Enemy.Name
        #Save Ship Locations and Status
        counter = 2
        for ship in Player.Ships.keys():
            for point in range(len(Player.Ships[ship].Points)):
                Player1.cell(row = counter, column = 2).value = Player.Ships[ship].Points[point]["Point"]
                Player1.cell(row = counter, column = 3).value = Player.Ships[ship].Points[point]["Status"]
                Player2.cell(row = counter, column = 2).value = Enemy.Ships[ship].Points[point]["Point"]
                Player2.cell(row = counter, column = 3).value = Enemy.Ships[ship].Points[point]["Status"]
                counter += 1
        #Save Player Guesses
        if len(Player.Guesses) > 0:
            for guess in Player.Guesses.keys():
                Player1.cell(row = counter, column = 1).value = "Guess"
                Player1.cell(row = counter, column = 2).value = guess
                Player1.cell(row = counter, column = 3).value = Player.Guesses[guess]
                counter += 1
        counter -= len(Player.Guesses)
        if len(Enemy.Guesses) > 0:
            for guess in Enemy.Guesses.keys():
                Player2.cell(row = counter, column = 1).value = "Guess"
                Player2.cell(row = counter, column = 2).value = guess
                Player2.cell(row = counter, column = 3).value = Enemy.Guesses[guess]
                counter += 1
        #Save Game State
        Game_State.cell(row = 2, column = 2).value = Game.Game
        Game_State.cell(row = 3, column = 2).value = Game.Turn
        Game_State.cell(row = 4, column = 2).value = Game.P1_Wins
        Game_State.cell(row = 5, column = 2).value = Game.P2_Wins
        Game_State.cell(row = 6, column = 2).value = Game.Active_Player
        Game_State.cell(row = 7, column = 2).value = Game.P1_Passcode
        Game_State.cell(row = 8, column = 2).value = Game.P2_Passcode
        Game_State.cell(row = 9, column = 2).value = True
        if Game.Online == False:
            wb.save(self.local_path)
        else:
            wb.save(self.online_path)

    def load_game(self, Player, Enemy, Game):
        if Game.Online == False:
            wb = load_workbook(self.local_path)
        else:
            wb = load_workbook(self.online_path)
        sheets = wb.sheetnames
        Player1 = wb[sheets[0]]
        Player2 = wb[sheets[1]]
        Game_State = wb[sheets[2]]
        #Load Names
        Player.Name = Player1.cell(row = 1, column = 1).value
        Enemy.Name = Player2.cell(row = 1, column = 1).value
        #Load Ship locations and status
        counter = 2
        for ship in Player.Ships.keys():
            for point in range(Player.Ships[ship].Length):
                #P1
                location = Player1.cell(row = counter, column = 2).value
                status = Player1.cell(row = counter, column = 3).value
                a = {"Point": location, "Status": status}
                key = Player.Grid.translate_key(location)
                Player.Grid.update_grid(key[0], key[1], status)
                Player.Ships[ship].Points.append(a)
                if status == "Sunk":
                    Player.Ships[ship].Sunk = True
                #P2
                location = Player2.cell(row = counter, column = 2).value
                status = Player2.cell(row = counter, column = 3).value
                a = {"Point": location, "Status": status}
                key = Enemy.Grid.translate_key(location)
                Enemy.Grid.update_grid(key[0], key[1], status)
                Enemy.Ships[ship].Points.append(a)
                if status == "Sunk":
                    Enemy.Ships[ship].Sunk = True
                counter += 1
        #Load Guesses
        #P1
        while Player1.cell(row = counter, column = 1).value:
            Guess = Player1.cell(row = counter, column = 2).value
            Status = Player1.cell(row = counter, column = 3).value
            Player.Guesses[Guess] = Status
            counter += 1
        #P2
        counter -= len(Player.Guesses)
        while Player2.cell(row = counter, column = 1).value:
            Guess = Player2.cell(row = counter, column = 2).value
            Status = Player2.cell(row = counter, column = 3).value
            Enemy.Guesses[Guess] = Status
            counter += 1
        for guess in Player.Guesses.keys():
            key = Player.Grid.translate_key(guess)
            Player.Guess_Grid.update_grid(key[0], key[1], Player.Guesses[guess])
            Enemy.Grid.update_grid(key[0], key[1], Player.Guesses[guess])
        for guess in Enemy.Guesses.keys():
            key = Enemy.Grid.translate_key(guess)
            Enemy.Guess_Grid.update_grid(key[0], key[1], Enemy.Guesses[guess])
            Player.Grid.update_grid(key[0], key[1], Enemy.Guesses[guess])
        #Load Game State
        Game.Game = Game_State.cell(row = 2, column = 2).value
        Game.Turn = Game_State.cell(row = 3, column = 2).value
        Game.P1_Wins = Game_State.cell(row = 4, column = 2).value
        Game.P2_Wins = Game_State.cell(row = 5, column = 2).value
        Game.Active_Player = Game_State.cell(row = 6, column = 2).value
        Game.P1_Passcode = Game_State.cell(row = 7, column = 2).value
        Game.P2_Passcode = Game_State.cell(row = 8, column = 2).value

    def delete_data(self, Online_or_Local):
        if Online_or_Local == "Online":
            wb = load_workbook(self.online_path)
        else:
            wb = load_workbook(self.local_path)
        sheets = wb.sheetnames
        Player1 = wb[sheets[0]]
        Player2 = wb[sheets[1]]
        Game_State = wb[sheets[2]]
        #Clear Save File
        for row in Player1['B2:C18']:
            for cell in row:
                cell.value = None
        for row in Player1['A19:C119']:
            for cell in row:
                cell.value = None
        for row in Player2['B2:C18']:
            for cell in row:
                cell.value = None
        for row in Player2['A19:C119']:
            for cell in row:
                cell.value = None
        for row in Game_State['B2:B9']:
            for cell in row:
                cell.value = None
        if Online_or_Local == "Online":
            wb.save(self.online_path)
        else:
            wb.save(self.local_path)
